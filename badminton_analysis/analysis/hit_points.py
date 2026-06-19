import csv
from typing import Dict, Tuple, List
import numpy as np
import cv2
from ..court.mapper import CourtMapper
import openpyxl
import os
from PIL import Image
import logging
from PIL import ImageDraw, ImageFont
import time
logger = logging.getLogger(__name__)

class ShuttlecockAnalyzer:
    def __init__(self, detect_csv_path, video_path, roi_corners, corners, player_1_hand, player_2_hand, player_region, max_distance: int = 180):
        self.file_path = detect_csv_path
        self.video_path = video_path
        self.image_save_path = os.path.join(os.path.dirname(self.file_path), "hit_points")
        os.makedirs(self.image_save_path, exist_ok=True)
        self.template_path = "templates/Statistical templates.xlsx"
        self.max_distance = max_distance
        self.min_y = roi_corners[0][1] - 100
        self.roi_corners = roi_corners
        self.player_1_hand = player_1_hand
        self.player_2_hand = player_2_hand
        self.player_1_skill_dict = {}
        self.player_2_skill_dict = {}
        self.court_mapper = CourtMapper(corners)
        # 定义技术映射
        self.player_2_skill_mapping = []
        workbook = openpyxl.load_workbook('templates/Technical-name-map.xlsx')
        sheet = workbook['Sheet1'] 
        for row in sheet.iter_rows(min_row=2, values_only=True, max_col=5):
            self.player_2_skill_mapping.append(row)
        print("start analyzing")

    def extract_shuttlecock_coordinates(self):
        coordinates_list = []
        current_dict = {}
        last_frame = None

        with open(self.file_path, 'r', newline='') as csvfile:
            reader = csv.reader(csvfile)
            next(reader)  # Skip header row
            for row in reader:
                frame = int(row[0])
                ball_x = int(row[11])
                ball_y = int(row[12])
                upper_court_x = float(row[3])
                upper_court_y = float(row[4])
                lower_court_x = float(row[8])
                lower_court_y = float(row[9])
                if self.player_1_hand == 'right':
                    upper_hand_x = float(row[15])
                    upper_hand_y = float(row[16])
                else:
                    upper_hand_x = float(row[13])
                    upper_hand_y = float(row[14])
                if self.player_2_hand == 'right':
                    lower_hand_x = float(row[19])
                    lower_hand_y = float(row[20])
                else:
                    lower_hand_x = float(row[17])
                    lower_hand_y = float(row[18])
                detect_frame_count = int(row[21])
                if last_frame is not None and frame - last_frame > 100:
                    # 将比赛分段获取
                    if len(current_dict) > 150:
                        coordinates_list.append(current_dict)
                    current_dict = {}

                current_dict[frame] = (ball_x, ball_y, upper_court_x, upper_court_y, lower_court_x, lower_court_y,
                                       upper_hand_x, upper_hand_y, lower_hand_x, lower_hand_y, detect_frame_count)
                last_frame = frame

        # 将最后一个字典保存
        if len(current_dict) > 150:
            coordinates_list.append(current_dict)

        return coordinates_list

    @staticmethod
    def distance(p1: Tuple[int, int], p2: Tuple[int, int]) -> float:
        return np.linalg.norm(np.array(p1) - np.array(p2))

    def filter_coordinates(self, coordinates):
        """
        过滤并清理羽毛球轨迹坐标数据
        
        Args:
            coordinates (dict): 包含帧索引和坐标的字典
                              格式: {frame_idx: (x, y, ...), ...}
        
        Returns:
            dict: 过滤后的坐标字典
            
        过滤规则:
        1. 移除连续重复的坐标点
        2. 移除超出合理高度范围的点
        3. 移除与前后帧距离异常的点
        """
        if not coordinates:
            return {}
        try:
            # 定义零坐标，但最后一个元素为保持不变      
            ZERO_COORD = (0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            MIN_Y_THRESHOLD = 100
            
            def is_valid_position(pos):
                return pos[:2] != (0, 0)
                
            def get_valid_position(positions, index, direction):
                """获取有效的前/后帧位置"""
                pos = positions[index + direction]
                if not is_valid_position(pos):
                    pos = positions[index + 2*direction]
                return pos
            
            avg_distances = {}
            first_index = list(coordinates.keys())[0]
            filtered_data = coordinates.copy()
            
            for index in range(2, len(coordinates) - 2):
                i = index + first_index
                current_pos = coordinates[i]
                
                # 过滤连续重复帧
                if current_pos[:2] == coordinates[i + 1][:2]:
                    filtered_data[i] = ZERO_COORD
                    continue
                    
                # 当前帧的球位置有效，且当前帧球的位置在很高的地方，则把当前帧的球位置置零
                # 没必要：当前帧的球位置在2运动员手的下方且运动员手的坐标有效 current_pos[1] > current_pos[-2] and current_pos[-3:-1] != self.roi_corners[0] 
                if (is_valid_position(current_pos) and 
                    (current_pos[1] < MIN_Y_THRESHOLD)):
                    filtered_data[i] = ZERO_COORD
                    continue
                
                # 获取前后帧位置
                prev_pos = get_valid_position(coordinates, i, -1)
                next_pos = get_valid_position(coordinates, i, 1)
                
                # 距离过滤
                if all(is_valid_position(pos) for pos in [current_pos, prev_pos, next_pos]):
                    avg_dist = (self.distance(current_pos[:2], next_pos[:2]) +
                               self.distance(current_pos[:2], prev_pos[:2])) / 2
                    avg_distances[i] = avg_dist
            return filtered_data
        except Exception as e:
            logger.error(f"Error filtering coordinates: {e}")
            return coordinates  # 返回原始数据作为fallback

    # 找击球点
    def find_potential_hits(self, filtered_data):
        potential_hits = []
        coordinates = list(filtered_data.items())
        # 找先向下后向上的帧
        for i in range(2, len(coordinates) - 2):
            frame, pos = coordinates[i]
            pre_frame, prev_pos = coordinates[i - 1]
            if prev_pos[:2] == (0, 0):
                pre_frame, prev_pos = coordinates[i - 2]
            fore_frame, fore_pos = coordinates[i + 1]
            if fore_pos[:2] == (0, 0):
                fore_frame, fore_pos = coordinates[i + 2]
            if pos[:2] == (0, 0) or pos[:2] == prev_pos[:2] or pos[:2] == fore_pos[:2]:
                continue
            if (pos[1] > fore_pos[1] and pos[1] > prev_pos[1] and prev_pos[1] != 0 and fore_pos[1] != 0):
                potential_hits.append((frame, pos))

        # 每15帧保留一个最低点
        filtered_hits = []
        i = 0
        while i < len(potential_hits):
            current_hit = potential_hits[i]
            j = i + 1
            while j < len(potential_hits) and potential_hits[j][0] - current_hit[0] <= 15:
                if potential_hits[j][1][1] > current_hit[1][1]:
                    current_hit = potential_hits[j]
                j += 1
            filtered_hits.append(current_hit)
            i = j

        # 把击球点分为前后两个人
        upper_potential_hits = []
        for hit in filtered_hits:
            if hit[1][6:8] != self.roi_corners[0] and hit[1][8:10] != self.roi_corners[0] \
                    and self.distance(hit[1][:2], hit[1][6:8]) < self.distance(hit[1][:2], hit[1][8:10]):
                upper_potential_hits.append(hit)
        lower_filtered_hits_set = set(filtered_hits)
        upper_potential_hits_set = set(upper_potential_hits)

        lower_filtered_hits_final = list(lower_filtered_hits_set - upper_potential_hits_set)
        lower_filtered_hits_final = sorted(lower_filtered_hits_final)

        print('lower_filtered_hits_final: ', lower_filtered_hits_final)
        return lower_filtered_hits_final, upper_potential_hits

    def find_miss_upper_hits(self, upper_hits_need_to_find, filtered_data):
        find_upper_hits = []
        for start, end in upper_hits_need_to_find:
            # 初始化为整数帧号
            nearest_frame = int((end + start) / 2)
            nearest_upper_hand_and_ball_distance = 200
            for i in range(start, end):
                if i in filtered_data and filtered_data[i][:2] != (0, 0) and filtered_data[i][6:8] != self.roi_corners[0]:
                    upper_hand_and_ball_distance = self.distance(filtered_data[i][:2], filtered_data[i][6:8])
                    if upper_hand_and_ball_distance < nearest_upper_hand_and_ball_distance:
                        nearest_upper_hand_and_ball_distance = upper_hand_and_ball_distance
                        nearest_frame = i
            # 确保nearest_frame存在于filtered_data中
            if nearest_frame in filtered_data:
                find_upper_hits.append((nearest_frame, filtered_data[nearest_frame]))
            else:
                # 如果计算出的nearest_frame不在filtered_data中，寻找最近的有效帧
                valid_frames = [f for f in filtered_data.keys() if start <= f <= end]
                if valid_frames:
                    # 找到最接近nearest_frame的有效帧
                    closest_frame = min(valid_frames, key=lambda x: abs(x - nearest_frame))
                    find_upper_hits.append((closest_frame, filtered_data[closest_frame]))
        return find_upper_hits

    def find_miss_lower_hits(self, lower_hits_need_to_find, filtered_data):
        find_lower_hits = []
        for start, end in lower_hits_need_to_find:
            # 初始化为整数帧号
            lowest_ball_frame = int((end + start) / 2)
            lowest_ball_y = 100
            for i in range(start + 5, end):
                if i in filtered_data and filtered_data[i][1] > lowest_ball_y:
                    lowest_ball_y = filtered_data[i][1]
                    lowest_ball_frame = i
            # 确保lowest_ball_frame存在于filtered_data中
            if lowest_ball_frame in filtered_data:
                find_lower_hits.append((lowest_ball_frame, filtered_data[lowest_ball_frame]))
            else:
                # 如果计算出的lowest_ball_frame不在filtered_data中，寻找最近的有效帧
                valid_frames = [f for f in filtered_data.keys() if start <= f <= end]
                if valid_frames:
                    # 找到最接近lowest_ball_frame的有效帧
                    closest_frame = min(valid_frames, key=lambda x: abs(x - lowest_ball_frame))
                    find_lower_hits.append((closest_frame, filtered_data[closest_frame]))
        return find_lower_hits

    # 找击球对
    def find_hits_pair(self, upper_hits_raw, lower_hits_raw, filtered_data):
        upper_hits = [upper_hit[0] for upper_hit in upper_hits_raw]
        lower_hits = [lower_hit[0] for lower_hit in lower_hits_raw]
        # 将player1找错的击球点挑出来，放到player2里去
        for i in range(1, len(upper_hits) - 1):
            if upper_hits[i + 1] - upper_hits[i] < 50 and upper_hits[i] - upper_hits[i - 1] < 50:
                missed_lower_hit = upper_hits_raw.pop(i)
                lower_hits_raw.append(missed_lower_hit)
        upper_hits = sorted([upper_hit[0] for upper_hit in upper_hits_raw])
        lower_hits = sorted([lower_hit[0] for lower_hit in lower_hits_raw])
        # print(upper_hits)
        # print(lower_hits)

        # 找player1没找到的击球点，算球离player1手最近的点
        upper_hits_need_to_find = []
        for i in range(len(lower_hits) - 1):
            flag = False
            for upper_hit in upper_hits:
                if upper_hit > lower_hits[i] and upper_hit < lower_hits[i + 1]:
                    flag = True
            if not flag:
                upper_hits_need_to_find.append((lower_hits[i], lower_hits[i + 1]))
        find_upper_hits = self.find_miss_upper_hits(upper_hits_need_to_find, filtered_data=filtered_data)

        upper_hits_all = sorted(find_upper_hits + upper_hits_raw)
        upper_hits_frame = [upper_hit[0] for upper_hit in upper_hits_all]

        # 找player2没找到的击球点，算最低时刻的
        lower_hits_need_to_find = []
        for i in range(len(upper_hits_frame) - 1):
            flag = False
            for lower_hit in lower_hits:
                if lower_hit > upper_hits_frame[i] and lower_hit < upper_hits_frame[i + 1]:
                    flag = True
            if not flag:
                lower_hits_need_to_find.append((upper_hits_frame[i], upper_hits_frame[i + 1]))
        # print(lower_hits_need_to_find)
        find_lower_hits = self.find_miss_lower_hits(lower_hits_need_to_find, filtered_data=filtered_data)
        lower_hits_all = sorted(find_lower_hits + lower_hits_raw)
        lower_hits_frame = [lower_hit[0] for lower_hit in lower_hits_all]
        # print(lower_hits_frame)

        # 找到upper_hits_frame中第一个比lower_hits_frame第一个元素大的元素的索引
        lower2upper_start_index = next(i for i, x in enumerate(upper_hits_frame) if x > lower_hits_frame[0])

        # 从找到的索引开始配对两个列表
        lower2upper = list(zip(lower_hits_frame, upper_hits_frame[lower2upper_start_index:]))

        # 找到lower_hits_frame中第一个比upper_hits_frame第一个元素大的元素的索引
        upper2lower_start_index = next(i for i, x in enumerate(lower_hits_frame) if x > upper_hits_frame[0])

        # 从找到的索引开始配对两个列表
        upper2lower = list(zip(upper_hits_frame, lower_hits_frame[upper2lower_start_index:]))
        print(lower2upper)
        return lower2upper, upper2lower

    def find_player_zone(self, player_position):
        x, y = player_position
        if x < 2.03:
            if y < 8.04 and y >= 6.7:
                return '1'
            elif y < 9.38 and y >= 8.04:
                return '4'
            elif y >= 9.38 and y < 10.72:
                return '7'
            elif y >= 10.72 and y < 12.06:
                return '10'
            elif y >= 12.06 and y < 13.4:
                return '13'
            elif y < 6.7 and y >= 5.36:
                return 'M'
            elif y >= 4.02 and y < 5.36:
                return 'J'
            elif y >= 2.68 and y < 4.02:
                return 'G'
            elif y >= 1.34 and y < 2.68:
                return 'D'
            elif y < 1.34:
                return 'A'
        elif x >= 2.03 and x < 4.06:
            if y < 8.04 and y >= 6.7:
                return '2'
            elif y < 9.38 and y >= 8.04:
                return '5'
            elif y >= 9.38 and y < 10.72:
                return '8'
            elif y >= 10.72 and y < 12.06:
                return '11'
            elif y >= 12.06 and y < 13.4:
                return '14'
            elif y < 6.7 and y >= 5.36:
                return 'N'
            elif y >= 4.02 and y < 5.36:
                return 'K'
            elif y >= 2.68 and y < 4.02:
                return 'H'
            elif y >= 1.34 and y < 2.68:
                return 'E'
            elif y < 1.34:
                return 'B'
        elif x >= 4.06:
            if y < 8.04 and y >= 6.7:
                return '3'
            elif y < 9.38 and y >= 8.04:
                return '6'
            elif y >= 9.38 and y < 10.72:
                return '9'
            elif y >= 10.72 and y < 12.06:
                return '12'
            elif y >= 12.06 and y < 13.4:
                return '15'
            elif y < 6.7 and y >= 5.36:
                return 'O'
            elif y >= 4.02 and y < 5.36:
                return 'L'
            elif y >= 2.68 and y < 4.02:
                return 'I'
            elif y >= 1.34 and y < 2.68:
                return 'F'
            elif y < 1.34:
                return 'C'
        return '0'

    def find_player_height(self, player_position, ball_position):
        image_player_position = self.court_mapper.court_to_image(player_position)
        player_height = "中"
        ball_height = image_player_position[1] - ball_position[1]
        if ball_height < 150:
            player_height = "低"
        elif ball_height >= 150 and ball_height < 200:
            player_height = "中"
        elif ball_height >= 200:
            player_height = "高"

        return player_height

    def find_player_2_skill_name(self, player_1_height, player_2_height, player_1_zone, player_2_zone):
        # 判断技术
        final_skill = ''
        for skill, zone_low, zone_up, hit_height, flight_height in self.player_2_skill_mapping:
            zone_low = str(zone_low).split('.')
            if player_1_zone in zone_up and player_2_zone in zone_low \
                    and player_2_height in hit_height and player_1_height in flight_height:
                if self.player_2_hand == 'right' and player_2_zone in ['1', '4', '7', '10', '13']:
                    final_skill = '反手_' + skill
                elif self.player_2_hand == 'left' and player_2_zone in ['3', '6', '9', '12', '15']:
                    final_skill = '反手_' + skill
                elif player_2_zone in ['2', '5', '8', '11', '14']:
                    final_skill = '中路_' + skill
                else:
                    final_skill = "正手_" + skill
                if final_skill not in self.player_2_skill_dict:
                    self.player_2_skill_dict[final_skill] = 1
                else:
                    self.player_2_skill_dict[final_skill] += 1
                break
        return final_skill

    def extract_and_stitch_frames(self, video_path, frame1, frame2, output_path, skill):
        # 打开视频文件
        cap = cv2.VideoCapture(video_path)

        # 检查视频是否成功打开
        if not cap.isOpened():
            print("Error: 无法打开视频文件")
            return

        # 获取视频的总帧数
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

        # 检查指定的帧是否有效
        if frame1 >= total_frames or frame2 >= total_frames or frame1 < 0 or frame2 < 0:
            print("Error: 指定的帧不在视频范围内")
            cap.release()
            return

        # 提取第一帧
        cap.set(cv2.CAP_PROP_POS_FRAMES, frame1)
        ret, img1 = cap.read()
        if not ret:
            print("Error: 无法读取第一帧")
            cap.release()
            return
        fps = cap.get(cv2.CAP_PROP_FPS)

        # 提取第二帧
        cap.set(cv2.CAP_PROP_POS_FRAMES, frame2)
        ret, img2 = cap.read()
        if not ret:
            print("Error: 无法读取第二帧")
            cap.release()
            return

        # 垂直拼接两帧
        stitched_image = np.vstack((img1, img2))

        # # 计算时间（秒）
        # total_seconds = frame1 / fps

        # # 转换为小时、分钟、秒
        # hours = int(total_seconds // 3600)
        # minutes = int((total_seconds % 3600) // 60)
        # seconds = int(total_seconds % 60)
        # time = f"{hours:02d}:{minutes:02d}:{seconds:02d}"

        # # 计算文字位置
        # height, width = stitched_image.shape[:2]
        # text_positions = [
        #     (width/2-100, height/2),
        #     (width/2-100, height/2 + 100),
        #     (width/2-100, height/2 + 200),
        #     (width/2-100, height/2 + 300),
        # ]
        # # 在拼接图片上添加文字
        # cv2.putText(stitched_image, time, tuple(map(int, (width/2-100, height/2-100))), cv2.FONT_HERSHEY_SIMPLEX, 3, (255, 255, 255), 5)
        # for (text, position) in zip(text_vars, text_positions):
        #     cv2.putText(stitched_image, text, tuple(map(int, position)),  cv2.FONT_HERSHEY_SIMPLEX, 3, (255, 255, 255), 5)
        
        # 检查目标文件是否已经存在
        if os.path.exists(output_path):
            # 如果文件已存在，直接删除它
            try:
                os.remove(output_path)
                print(f"已删除已存在的文件: {output_path}")
            except Exception as e:
                print(f"删除已存在文件失败: {e}")
                # 创建一个新的唯一文件名
                output_dir = os.path.dirname(output_path)
                base_name = os.path.basename(output_path)
                name_part, ext = os.path.splitext(base_name)
                output_path = os.path.join(output_dir, f"{name_part}_new{ext}")
        
        # 直接保存拼接后的图片到最终路径
        try:
            # 将OpenCV图像转换为PIL图像以便支持中文路径
            img_pil = Image.fromarray(cv2.cvtColor(stitched_image, cv2.COLOR_BGR2RGB))
            img_pil.save(output_path)
            print(f"拼接图片已保存至: {output_path}")
        except Exception as e:
            print(f"保存图片失败: {e}")
            # 尝试使用备用方法保存
            try:
                import uuid
                temp_filename = f"temp_image_{uuid.uuid4().hex}.png"
                cv2.imwrite(temp_filename, stitched_image)
                os.rename(temp_filename, output_path)
                print(f"使用备用方法保存图片至: {output_path}")
            except Exception as e2:
                print(f"备用保存方法也失败: {e2}")

        # 释放视频捕获对象
        cap.release()

    def extract_and_create_gif(self, video_path, frame1, frame2, output_path, skill):
        start_time = time.time()
        # 打开视频
        cap = cv2.VideoCapture(video_path)

        # 检查视频是否成功打开
        if not cap.isOpened():
            print("Error: 无法打开视频文件")
            return

        # 获取视频的总帧数
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

        # 检查指定的帧是否有效
        if frame1 >= total_frames or frame2 >= total_frames or frame1 < 0 or frame2 < 0:
            print("Error: 指定的帧不在视频范围内")
            cap.release()
            return

        # 创建帧列表
        frames = []
        fps = cap.get(cv2.CAP_PROP_FPS)

        # 计算时间（秒）
        total_seconds = frame1 / fps
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        seconds = int(total_seconds % 60)
        gif_time = f"{hours:02d}:{minutes:02d}:{seconds:02d}"

        # 提取帧范围内的所有帧
        for frame_idx in range(frame1, frame2 + 1):
            cap.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)
            ret, frame = cap.read()
            if not ret:
                print(f"Error: 无法读取帧 {frame_idx}")
                continue

            # 转换为PIL Image以处理中文
            frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            pil_image = Image.fromarray(frame)
            draw = ImageDraw.Draw(pil_image)
            
            # 加载常见系统中文字体；找不到时降级为 Pillow 默认字体。
            font = None
            font_candidates = [
                "C:/Windows/Fonts/simhei.ttf",
                "C:/Windows/Fonts/msyh.ttc",
                "/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf",
                "/System/Library/Fonts/PingFang.ttc",
            ]
            for font_path in font_candidates:
                if os.path.exists(font_path):
                    try:
                        font = ImageFont.truetype(font_path, 32)
                        break
                    except OSError:
                        continue
            if font is None:
                font = ImageFont.load_default()

            # 添加黑色背景
            text_bbox = draw.textbbox((10, 10), gif_time, font=font)
            draw.rectangle([text_bbox[0]-5, text_bbox[1]-5, text_bbox[2]+5, text_bbox[3]+5], fill='black')
            text_bbox = draw.textbbox((10, 50), skill, font=font)
            draw.rectangle([text_bbox[0]-5, text_bbox[1]-5, text_bbox[2]+5, text_bbox[3]+5], fill='black')

            # 添加文字
            draw.text((10, 10), gif_time, font=font, fill='white')
            draw.text((10, 50), skill, font=font, fill='white')

            frames.append(pil_image)

        # 释放视频捕获对象
        cap.release()

        # 删除已存在的输出文件
        if os.path.exists(output_path):
            os.remove(output_path)

        # 保存为GIF
        frames[0].save(
            output_path,
            save_all=True,
            append_images=frames[1:],
            duration=1000/fps,  # 持续时间（毫秒）
            loop=0
        )
        end_time = time.time()
        print(f"GIF已保存至: {output_path}")
        print(f"耗时: {end_time - start_time}秒")
        

    def data_statistics(self, lower2upper, filtered_data):
        player_2_skill_list = []
        for pair in lower2upper:
            data_pair = (filtered_data[pair[0]], filtered_data[pair[1]])
            player_2_position = data_pair[0][4:6]
            player_1_position = data_pair[1][2:4]
            player_2_zone = self.find_player_zone(player_2_position)
            player_1_zone = self.find_player_zone(player_1_position)

            ball_2_position = data_pair[0][:2]
            ball_1_position = data_pair[1][:2]
            player_2_height = self.find_player_height(player_2_position, ball_2_position)
            player_1_height = self.find_player_height(player_1_position, ball_1_position)

            skill = self.find_player_2_skill_name(player_1_height, player_2_height, player_1_zone, player_2_zone)

            detect_frame_count_1 = data_pair[0][-1] 
            detect_frame_count_2 = data_pair[1][-1]
            print(detect_frame_count_1, detect_frame_count_2)
            output_path = os.path.join(self.image_save_path, str(pair[0])) + f'_{pair[1]}_{skill}.png'
            # self.extract_and_create_gif(video_path=self.video_path, frame1=detect_frame_count_1, frame2=detect_frame_count_2, output_path=output_path, skill=skill)
            self.extract_and_stitch_frames(self.video_path, detect_frame_count_1, detect_frame_count_2, output_path, skill)
            player_2_skill_list.append(skill)
        return player_2_skill_list

    def write2xlsx(self):
        # 定义模板CSV文件和输出CSV文件的路径
        template_file = self.template_path
        output_file = self.file_path.replace('detect', 'statistic')
        output_file = output_file[:-4] + '.xlsx'
        wb = openpyxl.load_workbook(template_file)

        # Select the active sheet
        sheet = wb.active
        player2_mapping = {
            # 正手
            '正手_搓球': 'C3',
            '正手_推球（直线）': 'C4',
            '正手_推球（斜线）': 'C5',
            '正手_推球（中路）': 'C6',
            '正手_勾球': 'C7',
            '正手_挑球（直线）': 'C8',
            '正手_挑球（斜线）': 'C9',
            '正手_挑球（中路）': 'C10',
            '正手_放网（直线）': 'C11',
            '正手_放网（中路）': 'C12',
            '正手_扑球（直线）': 'C13',
            '正手_扑球（斜线）': 'C14',
            '正手_扑球（中路）': 'C15',
            '正手_挡（直线）': 'C18',
            '正手_挡（斜线）': 'C19',
            '正手_挡（右）': 'C20',
            '正手_抽（直线）': 'C21',
            '正手_抽（斜线）': 'C22',
            '正手_抽（中）': 'C23',
            '正手_弹（直线）': 'C24',
            '正手_弹（斜线）': 'C25',
            '正手_弹（中路）': 'C26',
            '正手_高球（直线）': 'C29',
            '正手_高球（斜线）': 'C30',
            '正手_高球（中路）': 'C31',
            '正手_吊球（直线）': 'C32',
            '正手_吊球（斜线）': 'C33',
            '正手_吊球（中路）': 'C34',
            '正手_杀球（直线）': 'C35',
            '正手_杀球（斜线）': 'C36',
            '正手_杀球（中路）': 'C37',

            # 反手
            '反手_搓球': 'F3',
            '反手_推球（直线）': 'F4',
            '反手_推球（斜线）': 'F5',
            '反手_推球（中路）': 'F6',
            '反手_勾球': 'F7',
            '反手_挑球（直线）': 'F8',
            '反手_挑球（斜线）': 'F9',
            '反手_挑球（中路）': 'F10',
            '反手_放网（直线）': 'F11',
            '反手_放网（中路）': 'F12',
            '反手_扑球（直线）': 'F13',
            '反手_扑球（斜线）': 'F14',
            '反手_扑球（中路）': 'F15',
            '反手_挡（直线）': 'F18',
            '反手_挡（斜线）': 'F19',
            '反手_挡（右）': 'F20',
            '反手_抽（直线）': 'F21',
            '反手_抽（斜线）': 'F22',
            '反手_抽（中）': 'F23',
            '反手_弹（直线）': 'F24',
            '反手_弹（斜线）': 'F25',
            '反手_弹（中路）': 'F26',
            '反手_高球（直线）': 'F29',
            '反手_高球（斜线）': 'F30',
            '反手_高球（中路）': 'F31',
            '反手_吊球（直线）': 'F32',
            '反手_吊球（斜线）': 'F33',
            '反手_吊球（中路）': 'F34',
            '反手_杀球（直线）': 'F35',
            '反手_杀球（斜线）': 'F36',
            '反手_杀球（中路）': 'F37',

            # 中场
            '中路_搓球': 'I3',
            '中路_勾球（左）': 'I4',
            '中路_勾球（右）': 'I5',
            '中路_推球（左）': 'I6',
            '中路_推球（中）': 'I7',
            '中路_推球（右）': 'I8',
            '中路_扑球（左）': 'I9',
            '中路_扑球（中）': 'I10',
            '中路_扑球（右）': 'I11',
            '中路_挑球（左）': 'I12',
            '中路_挑球（中）': 'I13',
            '中路_挑球（右）': 'I14',
            '中路_放（左）': 'I15',
            '中路_放（中）': 'I16',
            '中路_放（右）': 'I17',
            '中路_挡（左）': 'I20',
            '中路_挡（中）': 'I21',
            '中路_挡（右）': 'I22',
            '中路_抽（中）': 'I23',
            '中路_抽（左）': 'I24',
            '中路_抽（右）': 'I25',
            '中路_弹（左）': 'I26',
            '中路_弹（中）': 'I27',
            '中路_弹（右）': 'I28',
            '中路_高球（左）': 'I31',
            '中路_高球（中）': 'I32',
            '中路_高球（右）': 'I33',
            '中路_吊球（左）': 'I34',
            '中路_吊球（中）': 'I35',
            '中路_吊球（右）': 'I36',
            '中路_杀球（左）': 'I37',
            '中路_杀球（中）': 'I38',
            '中路_杀球（右）': 'I39',
        }
        # Fill in the data
        for key, value in self.player_2_skill_dict.items():
            if key in player2_mapping:
                cell = sheet[player2_mapping[key]]
                cell.value = value
        # Save the workbook
        wb.save(output_file)

        print(f"数据已成功写入 {output_file}")

    def analyze(self):
        coordinates_list = self.extract_shuttlecock_coordinates()
        for coordinates in coordinates_list:
            filtered_data = self.filter_coordinates(coordinates)
            lower_hits, upper_hits = self.find_potential_hits(filtered_data)
            lower2upper, upper2lower = self.find_hits_pair(upper_hits, lower_hits, filtered_data)
            player_2_skill_list = self.data_statistics(lower2upper, filtered_data)
            print(player_2_skill_list)
            #把player_2_skill_list写入xlsx文件
            # with open('results/yolo_00002/player_2_skill_list.csv', 'w') as file:
            #     writer = csv.writer(file)
            #     writer.writerow(player_2_skill_list)
        print(self.player_2_skill_dict)
        self.write2xlsx()

# Example usage
if __name__ == "__main__":
    # 定义基础路径和视频名称
    base_path = 'results'
    # video_name = 'youtube3_clip1'  # 只需修改这个变量
    # video_name = 'youtube1_clip1'  # 只需修改这个变量
    video_name = '00003'  # 只需修改这个变量
    # video_name = 'bilibili-1'  # 只需修改这个变量
    # video_name = 'youtube2'  # 只需修改这个变量
    # video_name = 'youtube1'  # 只需修改这个变量
   
    # 构建完整路径
    result_folder = f'{base_path}/{video_name}'
    file_path = f'{result_folder}/detect_{video_name}.csv'
    video_path = f'{result_folder}/detect_{video_name}.mp4'
    court_annotation_path = f'{result_folder}/court_annotations.txt'

    # 读取场地标注数据
    with open(court_annotation_path, 'r') as file:
        data = file.read().replace('\n', '')
    corners, roi_corners = map(eval, [
        data.split('corners=')[1].split(']')[0] + ']', 
        data.split('roi_corners=')[1].split(']')[0] + ']'
    ])


    analyzer = ShuttlecockAnalyzer(detect_csv_path=file_path, video_path=video_path, roi_corners=roi_corners, corners=corners, player_1_hand='right', player_2_hand='right', player_region='upper')
    analyzer.analyze()
